import os
import sys

from PySide6.QtCore import Qt, Signal
from PySide6.QtWidgets import (
    QDialog,
    QFileDialog,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QPushButton,
    QVBoxLayout,
    QWidget,
)

from app.ui.theme import PRIMARY, TEXT_SECONDARY


class _ImportDialog(QDialog):
    import_ready = Signal(dict)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("导入会议计划")
        self.setFixedSize(380, 180)
        self.setWindowFlags(Qt.WindowType.Dialog)

        self._setup_ui()

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        title_label = QLabel("导入会议计划")
        title_label.setStyleSheet(
            f"color: {TEXT_SECONDARY}; font-size: 16px; font-weight: 600;"
        )
        layout.addWidget(title_label)

        desc_label = QLabel("通过Excel文件导入会议议程，文件名将作为会议名称")
        desc_label.setStyleSheet(
            f"color: {TEXT_SECONDARY}; font-size: 13px;"
        )
        desc_label.setWordWrap(True)
        layout.addWidget(desc_label)

        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(8)

        self._download_btn = QPushButton("下载模板")
        self._download_btn.setObjectName("secondaryBtn")
        self._download_btn.clicked.connect(self._on_download_template)
        btn_layout.addWidget(self._download_btn)

        self._import_btn = QPushButton("导入文件")
        self._import_btn.setObjectName("primaryBtn")
        self._import_btn.clicked.connect(self._on_import_file)
        btn_layout.addWidget(self._import_btn)

        self._cancel_btn = QPushButton("取消")
        self._cancel_btn.setObjectName("secondaryBtn")
        self._cancel_btn.clicked.connect(self.reject)
        btn_layout.addWidget(self._cancel_btn)

        layout.addLayout(btn_layout)

    def _get_template_path(self):
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, 'resources', 'templates', '会议议程导入模板.xlsx')
        return os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            'resources', 'templates', '会议议程导入模板.xlsx'
        )

    def _on_download_template(self):
        template_path = self._get_template_path()
        if not os.path.exists(template_path):
            QMessageBox.warning(self, "提示", "模板文件不存在")
            return

        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存模板",
            "会议议程导入模板.xlsx",
            "Excel文件 (*.xlsx)"
        )

        if save_path:
            if not save_path.endswith('.xlsx'):
                save_path += '.xlsx'
            try:
                import shutil
                shutil.copy(template_path, save_path)
                QMessageBox.information(self, "提示", "模板下载成功")
                self.accept()
            except Exception as e:
                QMessageBox.warning(self, "提示", f"下载失败：{str(e)}")

    def _on_import_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择Excel文件",
            "",
            "Excel文件 (*.xlsx)"
        )

        if not file_path:
            return

        try:
            data = self._parse_excel(file_path)
            if data:
                self.import_ready.emit(data)
                self.accept()
        except Exception as e:
            QMessageBox.warning(self, "导入失败", str(e))

    def _parse_excel(self, file_path):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise Exception("请先安装openpyxl库")

        wb = load_workbook(file_path)
        ws = wb.active

        if ws.max_row < 2:
            raise Exception("Excel文件中没有数据")

        headers = [cell.value for cell in ws[1]]
        if headers[0] != '议题' or headers[1] != '汇报时间（分钟）' or headers[2] != '讨论时间（分钟）':
            raise Exception("Excel格式不正确，请使用下载的模板")

        topics = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue

            topic_name = str(row[0]).strip()
            if not topic_name:
                continue

            try:
                presentation_minutes = int(row[1]) if row[1] else 10
                qa_minutes = int(row[2]) if row[2] else 5
            except (ValueError, TypeError):
                raise Exception(f"第{ws._current_row}行时间格式不正确")

            if presentation_minutes <= 0:
                presentation_minutes = 10
            if qa_minutes <= 0:
                qa_minutes = 5

            topics.append({
                "name": topic_name,
                "presentation_minutes": presentation_minutes,
                "qa_minutes": qa_minutes,
            })

        if not topics:
            raise Exception("Excel文件中没有有效议题")

        file_name = os.path.basename(file_path)
        meeting_name = os.path.splitext(file_name)[0]

        return {
            "name": meeting_name,
            "topics": topics,
        }

    def exec(self):
        return super().exec()