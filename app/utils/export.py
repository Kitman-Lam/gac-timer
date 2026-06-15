import csv
from io import StringIO

SOUND_WARNING = "warning"
SOUND_TIMEUP = "timeup"
SOUND_OVERTIME = "overtime"


def _format_seconds(seconds: float) -> str:
    total_seconds = int(seconds)
    minutes = total_seconds // 60
    secs = total_seconds % 60
    return f"{minutes}:{secs:02d}"


def export_to_csv(meeting_data: dict, file_path: str):
    output = StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "\u8bae\u9898\u540d\u79f0",
        "\u9636\u6bb5",
        "\u8ba1\u5212\u65f6\u957f(\u5206\u949f)",
        "\u5b9e\u9645\u7528\u65f6(\u5206:\u79d2)",
        "\u8d85\u65f6\u65f6\u957f(\u5206:\u79d2)",
        "\u5360\u6bd4",
    ])

    total_actual = meeting_data.get("total_actual_seconds", 0.0)

    for topic in meeting_data.get("topics", []):
        for phase_key, phase_label in [("presentation", "\u6c47\u62a5"), ("qa", "\u8ba8\u8bba")]:
            phase = topic.get(phase_key, {})
            planned = phase.get("planned_minutes", 0)
            actual = phase.get("actual_seconds", 0.0)
            overtime = phase.get("overtime_seconds", 0.0)

            if total_actual > 0:
                percentage = f"{actual / total_actual * 100:.1f}%"
            else:
                percentage = "0.0%"

            writer.writerow([
                topic.get("name", ""),
                phase_label,
                planned,
                _format_seconds(actual),
                _format_seconds(overtime),
                percentage,
            ])

    writer.writerow([
        "\u5408\u8ba1",
        "",
        meeting_data.get("total_planned_minutes", 0),
        _format_seconds(meeting_data.get("total_actual_seconds", 0.0)),
        _format_seconds(meeting_data.get("total_overtime_seconds", 0.0)),
        "100.0%",
    ])

    content = output.getvalue()
    output.close()

    with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
        f.write(content)


def export_to_excel(meeting_data: dict, file_path: str):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    meeting_name = meeting_data.get("meeting_name", "Meeting")
    ws.title = meeting_name[:31]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    summary_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    summary_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    headers = [
        "\u8bae\u9898\u540d\u79f0",
        "\u9636\u6bb5",
        "\u8ba1\u5212\u65f6\u957f(\u5206\u949f)",
        "\u5b9e\u9645\u7528\u65f6(\u5206:\u79d2)",
        "\u8d85\u65f6\u65f6\u957f(\u5206:\u79d2)",
        "\u5360\u6bd4",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    total_actual = meeting_data.get("total_actual_seconds", 0.0)
    row = 2

    for topic in meeting_data.get("topics", []):
        for phase_key, phase_label in [("presentation", "\u6c47\u62a5"), ("qa", "\u8ba8\u8bba")]:
            phase = topic.get(phase_key, {})
            planned = phase.get("planned_minutes", 0)
            actual = phase.get("actual_seconds", 0.0)
            overtime = phase.get("overtime_seconds", 0.0)

            if total_actual > 0:
                percentage = actual / total_actual
            else:
                percentage = 0.0

            ws.cell(row=row, column=1, value=topic.get("name", "")).alignment = left_align
            ws.cell(row=row, column=2, value=phase_label).alignment = center_align
            ws.cell(row=row, column=3, value=planned).alignment = center_align
            ws.cell(row=row, column=4, value=_format_seconds(actual)).alignment = center_align
            ws.cell(row=row, column=5, value=_format_seconds(overtime)).alignment = center_align
            pct_cell = ws.cell(row=row, column=6, value=percentage)
            pct_cell.number_format = "0.0%"
            pct_cell.alignment = center_align

            row += 1

    ws.cell(row=row, column=1, value="\u5408\u8ba1").alignment = left_align
    ws.cell(row=row, column=2, value="").alignment = center_align
    ws.cell(row=row, column=3, value=meeting_data.get("total_planned_minutes", 0)).alignment = center_align
    ws.cell(row=row, column=4, value=_format_seconds(meeting_data.get("total_actual_seconds", 0.0))).alignment = center_align
    ws.cell(row=row, column=5, value=_format_seconds(meeting_data.get("total_overtime_seconds", 0.0))).alignment = center_align
    total_pct_cell = ws.cell(row=row, column=6, value=1.0)
    total_pct_cell.number_format = "0.0%"
    total_pct_cell.alignment = center_align

    for col_idx in range(1, 7):
        cell = ws.cell(row=row, column=col_idx)
        cell.fill = summary_fill
        cell.font = summary_font

    for col_idx in range(1, 7):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                char_count = 0
                for ch in str(cell_value):
                    char_count += 2 if "\u4e00" <= ch <= "\u9fff" else 1
                max_length = max(max_length, char_count)
        ws.column_dimensions[col_letter].width = max(max_length + 4, 10)

    import os
    # 检查路径是否存在
    dir_path = os.path.dirname(file_path)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path)
    wb.save(file_path)
